import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Kalkulator HPP", page_icon="🧮", layout="centered")

# ============================================================
# STATE: draft laporan (list produk yang sudah "ditambahkan")
# ============================================================
if "draft" not in st.session_state:
    st.session_state.draft = []  # list of dict per produk

st.title("🧮 Kalkulator & Draft Laporan HPP")
st.caption("Hitung HPP per produk, kumpulkan ke draft, lalu export semua produk ke Excel")

st.divider()

# ============================================================
# 0. NAMA PRODUK
# ============================================================
st.subheader("🏷️ Nama Produk")
nama_produk = st.text_input("Nama produk", placeholder="Contoh: Kaos Polos Combed 24s")

st.divider()

# ============================================================
# 1. HARGA MODAL
# ============================================================
st.subheader("1️⃣ Harga Modal")
harga_modal = st.number_input(
    "Harga modal / bahan baku per produk (Rp)",
    min_value=0.0, value=0.0, step=500.0, format="%.0f"
)

# ============================================================
# 2. PACKAGING
# ============================================================
st.subheader("2️⃣ Biaya Packaging")
biaya_packaging = st.number_input(
    "Biaya packaging per produk (Rp)",
    min_value=0.0, value=0.0, step=500.0, format="%.0f"
)

# ============================================================
# 3. RENCANA HARGA JUAL
# ============================================================
st.subheader("3️⃣ Rencana Harga Jual Produk")
harga_jual = st.number_input(
    "Rencana harga jual produk (Rp)",
    min_value=0.0, value=0.0, step=500.0, format="%.0f"
)

# ============================================================
# 4. BIAYA SHOPEE (rincian nominal Rp, bukan persentase)
# ============================================================
st.subheader("4️⃣ Biaya Shopee")
st.caption(
    "Isi sesuai angka riil yang muncul di Rincian Pesanan / Rincian Penghasilan Seller Center kamu. "
    "Beberapa komponen ini opsional, tergantung program yang kamu ikuti."
)

biaya_admin_shopee = st.number_input(
    "Biaya Administrasi Shopee (Rp)",
    min_value=0.0, value=0.0, step=500.0, format="%.0f",
    help="Potongan utama per transaksi, besarannya tergantung kategori produk & status toko (Non-Star/Star/Star+)."
)
biaya_proses_pesanan = st.number_input(
    "Biaya Proses Pesanan (Rp)",
    min_value=0.0, value=1250.0, step=250.0, format="%.0f",
    help="Biaya tetap per transaksi yang berhasil diselesaikan (umumnya Rp1.250)."
)
biaya_layanan_gratis_ongkir = st.number_input(
    "Biaya Layanan Gratis Ongkir XTRA (Rp)",
    min_value=0.0, value=0.0, step=500.0, format="%.0f",
    help="Hanya berlaku kalau toko ikut program Gratis Ongkir XTRA (maks. sekitar Rp10.000/produk)."
)
biaya_layanan_promo_xtra = st.number_input(
    "Biaya Layanan Promo XTRA (Rp)",
    min_value=0.0, value=0.0, step=500.0, format="%.0f",
    help="Hanya berlaku kalau toko ikut program Promo XTRA (maks. sekitar Rp10.000/produk)."
)
ongkir_ditanggung_penjual = st.number_input(
    "Ongkir / Subsidi Ongkir Ditanggung Penjual (Rp)",
    min_value=0.0, value=0.0, step=500.0, format="%.0f",
    help="Kalau kamu ikut program subsidi ongkir tertentu yang membebankan sebagian ongkir ke penjual."
)
biaya_shopee_lainnya = st.number_input(
    "Biaya Shopee Lainnya (Rp)",
    min_value=0.0, value=0.0, step=500.0, format="%.0f",
    help="Contoh: biaya layanan Pre-Order, Cashback XTRA, komisi affiliate, dll — isi kalau relevan."
)

biaya_shopee = (
    biaya_admin_shopee + biaya_proses_pesanan + biaya_layanan_gratis_ongkir
    + biaya_layanan_promo_xtra + ongkir_ditanggung_penjual + biaya_shopee_lainnya
)
st.write(f"**Total Biaya Shopee: Rp {biaya_shopee:,.0f}**".replace(",", "."))

# ============================================================
# 4. BIAYA OPERASIONAL
# ============================================================
st.subheader("5️⃣ Biaya Operasional")
st.caption("Persentase dari harga jual (gaji, listrik, bensin, dll dijumlah jadi satu persentase)")

persen_operasional = st.number_input(
    "Total biaya operasional (%) dari harga jual",
    min_value=0.0, max_value=100.0, value=10.0, step=0.5
)
biaya_operasional_per_unit = harga_jual * (persen_operasional / 100)

st.divider()

# ============================================================
# HITUNG HPP
# ============================================================
hpp = harga_modal + biaya_packaging + biaya_shopee + biaya_operasional_per_unit
laba = harga_jual - hpp
margin_persen = (laba / harga_jual * 100) if harga_jual > 0 else 0

st.subheader("📊 Hasil Perhitungan")

rincian = {
    "Harga Modal": harga_modal,
    "Biaya Packaging": biaya_packaging,
    "Biaya Shopee (total)": biaya_shopee,
    "Biaya Operasional per Unit": biaya_operasional_per_unit,
}
for label, nilai in rincian.items():
    st.write(f"- {label}: **Rp {nilai:,.0f}**".replace(",", "."))

st.metric("💰 Total HPP per Produk", f"Rp {hpp:,.0f}".replace(",", "."))

if harga_jual > 0:
    colA, colB = st.columns(2)
    colA.metric("Laba per Produk", f"Rp {laba:,.0f}".replace(",", "."))
    colB.metric("Margin (%)", f"{margin_persen:.1f}%")

    if laba < 0:
        st.error("⚠️ Harga jual masih di bawah HPP, kamu rugi di harga ini!")
    elif margin_persen < 10:
        st.warning("Margin masih tipis, pertimbangkan naikkan harga jual atau efisiensi biaya.")
    else:
        st.success("Margin cukup sehat 👍")
else:
    st.info("Isi 'Harga jual produk' di atas untuk melihat estimasi laba & margin.")

st.divider()
target_margin = st.number_input("Target margin keuntungan (%)", min_value=0.0, value=20.0, step=1.0)
harga_jual_disarankan = hpp / (1 - target_margin / 100) if target_margin < 100 else 0
st.write(f"Rekomendasi harga jual untuk margin {target_margin:.0f}%: "
         f"**Rp {harga_jual_disarankan:,.0f}**".replace(",", "."))

st.divider()

# ============================================================
# TAMBAH KE DRAFT
# ============================================================
st.subheader("➕ Tambah ke Draft Laporan")

tambah_disabled = not nama_produk.strip()
if tambah_disabled:
    st.caption("Isi 'Nama produk' dulu di atas untuk bisa menambahkan ke draft.")

if st.button("➕ Tambah ke Draft", type="primary", disabled=tambah_disabled):
    st.session_state.draft.append({
        "Nama Produk": nama_produk.strip(),
        "Harga Modal": harga_modal,
        "Biaya Packaging": biaya_packaging,
        "Harga Jual": harga_jual,
        "Biaya Admin Shopee": biaya_admin_shopee,
        "Biaya Proses Pesanan": biaya_proses_pesanan,
        "Biaya Layanan Gratis Ongkir XTRA": biaya_layanan_gratis_ongkir,
        "Biaya Layanan Promo XTRA": biaya_layanan_promo_xtra,
        "Ongkir Ditanggung Penjual": ongkir_ditanggung_penjual,
        "Biaya Lainnya": biaya_shopee_lainnya,
        "% Biaya Operasional": persen_operasional,
    })
    st.success(f"'{nama_produk.strip()}' ditambahkan ke draft ✅")

# ============================================================
# DRAFT LAPORAN
# ============================================================
st.divider()
st.subheader(f"📋 Draft Laporan HPP ({len(st.session_state.draft)} produk)")

SHOPEE_COLS = [
    "Biaya Admin Shopee", "Biaya Proses Pesanan", "Biaya Layanan Gratis Ongkir XTRA",
    "Biaya Layanan Promo XTRA", "Ongkir Ditanggung Penjual", "Biaya Shopee Lainnya",
]

if not st.session_state.draft:
    st.info("Draft masih kosong. Isi kalkulator di atas lalu klik 'Tambah ke Draft'.")
else:
    df_preview = pd.DataFrame(st.session_state.draft).copy()
    df_preview["Total Biaya Shopee"] = df_preview[SHOPEE_COLS].sum(axis=1)
    df_preview["Biaya Operasional"] = df_preview["Harga Jual"] * df_preview["% Biaya Operasional"] / 100
    df_preview["Total HPP"] = (
        df_preview["Harga Modal"] + df_preview["Biaya Packaging"]
        + df_preview["Total Biaya Shopee"] + df_preview["Biaya Operasional"]
    )
    df_preview["Laba"] = df_preview["Harga Jual"] - df_preview["Total HPP"]
    df_preview["Margin %"] = (df_preview["Laba"] / df_preview["Harga Jual"] * 100).fillna(0)

    st.dataframe(
        df_preview[["Nama Produk", "Harga Modal", "Biaya Packaging", "Harga Jual",
                     "Total Biaya Shopee", "% Biaya Operasional", "Total HPP", "Laba", "Margin %"]],
        use_container_width=True, hide_index=True
    )

    colX, colY = st.columns([2, 1])
    with colX:
        hapus_index = st.selectbox(
            "Pilih produk yang mau dihapus dari draft",
            options=list(range(len(st.session_state.draft))),
            format_func=lambda i: st.session_state.draft[i]["Nama Produk"]
        )
    with colY:
        st.write("")
        st.write("")
        if st.button("🗑️ Hapus Produk Ini"):
            st.session_state.draft.pop(hapus_index)
            st.rerun()

    if st.button("🧹 Kosongkan Semua Draft"):
        st.session_state.draft = []
        st.rerun()

    st.divider()

    # ============================================================
    # EXPORT TO EXCEL (hanya kalau draft sudah "selesai")
    # ============================================================
    st.subheader("📥 Export ke Excel")
    selesai = st.checkbox("✅ Draft sudah lengkap dan siap di-export")

    if selesai:
        def build_excel(draft_rows):
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan HPP"

            FONT_NAME = "Arial"
            HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
            INPUT_FONT = Font(name=FONT_NAME, color="0000FF", size=11)
            FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=11)
            TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
            NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="808080")
            CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
            LEFT = Alignment(horizontal="left", vertical="center")
            thin = Side(style="thin", color="BFBFBF")
            BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

            N_COLS = 17  # total kolom

            ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
            ws["A1"] = "LAPORAN HPP PER PRODUK"
            ws["A1"].font = TITLE_FONT
            ws["A1"].alignment = LEFT

            ws.merge_cells(f"A2:{get_column_letter(N_COLS)}2")
            ws["A2"] = f"Diexport dari Kalkulator HPP — {datetime.now().strftime('%d %B %Y %H:%M')}"
            ws["A2"].font = NOTE_FONT
            ws["A2"].alignment = LEFT

            HEADER_ROW = 4
            headers = [
                "No", "Nama Produk", "Harga Modal (Rp)", "Biaya Packaging (Rp)", "Harga Jual (Rp)",
                "Biaya Admin Shopee (Rp)", "Biaya Proses Pesanan (Rp)",
                "Biaya Layanan Gratis Ongkir XTRA (Rp)", "Biaya Layanan Promo XTRA (Rp)",
                "Ongkir Ditanggung Penjual (Rp)", "Biaya Shopee Lainnya (Rp)",
                "Total Biaya Shopee (Rp)", "% Biaya Operasional", "Biaya Operasional (Rp)",
                "Total HPP (Rp)", "Laba per Produk (Rp)", "Margin (%)"
            ]
            for col, h in enumerate(headers, start=1):
                c = ws.cell(row=HEADER_ROW, column=col, value=h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = CENTER
                c.border = BORDER

            col_widths = [5, 22, 14, 14, 12, 14, 14, 16, 15, 15, 14, 15, 13, 15, 13, 15, 10]
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            first_data_row = HEADER_ROW + 1
            last_data_row = first_data_row + len(draft_rows) - 1

            # Kolom: 1 No,2 Nama,3 Modal,4 Packaging,5 HargaJual,
            # 6 AdminShopee,7 ProsesPesanan,8 GratisOngkirXtra,9 PromoXtra,10 OngkirDitanggung,11 ShopeeLainnya,
            # 12 TotalShopee(formula),13 %Operasional,14 BiayaOperasional(formula),
            # 15 TotalHPP(formula),16 Laba(formula),17 Margin(formula)
            for i, row in enumerate(draft_rows):
                r = first_data_row + i
                ws.cell(row=r, column=1, value=i + 1)
                ws.cell(row=r, column=2, value=row["Nama Produk"])
                ws.cell(row=r, column=3, value=row["Harga Modal"])
                ws.cell(row=r, column=4, value=row["Biaya Packaging"])
                ws.cell(row=r, column=5, value=row["Harga Jual"])
                ws.cell(row=r, column=6, value=row["Biaya Admin Shopee"])
                ws.cell(row=r, column=7, value=row["Biaya Proses Pesanan"])
                ws.cell(row=r, column=8, value=row["Biaya Layanan Gratis Ongkir XTRA"])
                ws.cell(row=r, column=9, value=row["Biaya Layanan Promo XTRA"])
                ws.cell(row=r, column=10, value=row["Ongkir Ditanggung Penjual"])
                ws.cell(row=r, column=11, value=row["Biaya Shopee Lainnya"])
                ws.cell(row=r, column=12, value=f"=SUM(F{r}:K{r})")
                ws.cell(row=r, column=13, value=row["% Biaya Operasional"] / 100)
                ws.cell(row=r, column=14, value=f"=IFERROR(E{r}*M{r},0)")
                ws.cell(row=r, column=15, value=f"=C{r}+D{r}+L{r}+N{r}")
                ws.cell(row=r, column=16, value=f"=E{r}-O{r}")
                ws.cell(row=r, column=17, value=f"=IFERROR(P{r}/E{r},0)")

                for col in range(1, N_COLS + 1):
                    cell = ws.cell(row=r, column=col)
                    cell.border = BORDER
                    if col == 1:
                        cell.alignment = CENTER
                    if col == 2:
                        cell.alignment = LEFT
                    if col in (2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 13):
                        cell.font = INPUT_FONT
                    else:
                        cell.font = FORMULA_FONT
                    if col in (3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16):
                        cell.number_format = '#,##0;(#,##0);"-"'
                    if col in (13, 17):
                        cell.number_format = '0.0%'

            total_row = last_data_row + 2
            ws.merge_cells(f"A{total_row}:B{total_row}")
            ws.cell(row=total_row, column=1,
                    value=f"TOTAL / RATA-RATA ({len(draft_rows)} PRODUK)").font = Font(name=FONT_NAME, bold=True)

            sum_cols = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16]
            avg_cols = [13, 17]
            for col in sum_cols:
                letter = get_column_letter(col)
                ws.cell(row=total_row, column=col, value=f"=SUM({letter}{first_data_row}:{letter}{last_data_row})")
            for col in avg_cols:
                letter = get_column_letter(col)
                ws.cell(row=total_row, column=col,
                        value=f"=IFERROR(AVERAGE({letter}{first_data_row}:{letter}{last_data_row}),0)")

            for col in range(1, N_COLS + 1):
                cell = ws.cell(row=total_row, column=col)
                cell.font = Font(name=FONT_NAME, bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.border = BORDER
                if col in sum_cols:
                    cell.number_format = '#,##0;(#,##0);"-"'
                if col in avg_cols:
                    cell.number_format = '0.0%'

            ws.freeze_panes = f"A{first_data_row}"

            note_row = total_row + 2
            ws.merge_cells(f"A{note_row}:{get_column_letter(N_COLS)}{note_row}")
            ws.cell(row=note_row, column=1,
                    value="Catatan: sel teks biru = input, sel teks hitam = rumus otomatis "
                          "(Total Biaya Shopee, Biaya Operasional, Total HPP, Laba, Margin).")
            ws.cell(row=note_row, column=1).font = NOTE_FONT

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer

        excel_buffer = build_excel(st.session_state.draft)
        st.download_button(
            label="📥 Export to Excel",
            data=excel_buffer,
            file_name=f"laporan_hpp_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
    else:
        st.caption("Centang dulu 'Draft sudah lengkap dan siap di-export' untuk memunculkan tombol export.")
